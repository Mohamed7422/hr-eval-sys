# app/services/hierarchy_importer.py
from pathlib import Path
from io import TextIOWrapper
import csv
from typing import List, Dict, Any

try:
    import openpyxl  # for .xlsx support
except ImportError:
    openpyxl = None

from django.db import transaction
from django.db.models import QuerySet

from evaluation_app.models import Company, Department, SubDepartment, Section, SubSection  # adjust import to your app


# ---------- Public API ----------

def parse_hierarchy_rows(request) -> List[Dict[str, Any]]:
    """
    Return list[dict] from either JSON array or multipart CSV/XLSX uploaded under the 'file' key.
    """
    if "file" in request.FILES:
        f = request.FILES["file"]
        suffix = Path(f.name).suffix.lower()
        if suffix == ".csv":
            text = TextIOWrapper(f.file, encoding="utf-8", newline="")
            reader = csv.DictReader(text)
            rows = list(reader)
            if not rows:
                raise ValueError("CSV appears empty.")
            return rows

        if suffix in {".xlsx", ".xls"}:
            if openpyxl is None:
                raise ValueError("XLSX import requires 'openpyxl'. Install it or upload CSV.")
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in r):
                    continue
                rows.append({headers[i]: r[i] for i in range(len(headers))})
            if not rows:
                raise ValueError("XLSX sheet appears empty.")
            return rows

        raise ValueError("Unsupported file type. Upload CSV or XLSX.")

    # JSON
    data = request.data
    if not isinstance(data, list):
        raise ValueError('Expected a JSON array or upload a file as "file".')
    return data


def import_hierarchy_paths(rows: List[Dict[str, Any]], *, dry_run: bool = False) -> Dict[str, Any]:
    """
    Import paths shaped like:
        company -> department -> sub_department -> section -> sub_section

    - Companies MUST already exist (exact name).
    - Creates only the missing nodes; no renames or updates. Managers and employee_count are left null/0 at creation.
    - Empty cell means that level is not present on that row.

    Returns a dict suitable to be sent directly in a DRF Response.
    """
    cleaned = _normalize_and_select_headers(rows)  # adds "__row" index
    errors, company_names = _basic_validate(cleaned)
    if errors:
        return {"status": "invalid", "errors": errors}

    # Companies must exist
    companies = {c.name: c for c in Company.objects.filter(name__in=company_names)}
    missing = [n for n in company_names if n not in companies]
    if missing:
        errs = []
        for row in cleaned:
            if row["company"] in missing:
                errs.append({"row": row["__row"], "errors": {"company": [f'Company "{row["company"]}" does not exist. Import companies first.']}})
        return {"status": "invalid", "errors": errs}

    # Build unique keys for lookups to avoid duplicates
    dept_keys, subdept_keys, section_keys, subsection_keys = _collect_unique_keys(cleaned)

    # Prefetch existing
    dept_map = _fetch_departments(company_names, dept_keys)
    subdept_map = _fetch_subdepartments(company_names, subdept_keys)
    section_map = _fetch_sections(company_names, section_keys)
    subsection_map = _fetch_subsections(company_names, subsection_keys)

    # Decide what to create
    depts_to_create, subdepts_to_create, sections_to_create, subsections_to_create = [], [], [], []
    for r in cleaned:
        company = companies[r["company"]]
        d = r["department"]; sd = r["sub_department"]; s = r["section"]; ss = r["sub_section"]

        dept_obj = None
        if d:
            dkey = (company.name, d)
            dept_obj = dept_map.get(dkey)
            if not dept_obj:
                dept_obj = Department(name=d, company=company, employee_count=0, manager=None)
                depts_to_create.append(dept_obj)
                dept_map[dkey] = dept_obj  # temp bind

        subdept_obj = None
        if d and sd:
            sdkey = (company.name, d, sd)
            subdept_obj = subdept_map.get(sdkey)
            if not subdept_obj:
                subdept_obj = SubDepartment(name=sd, department=dept_obj, employee_count=0, manager=None)
                subdepts_to_create.append(subdept_obj)
                subdept_map[sdkey] = subdept_obj

        section_obj = None
        if d and sd and s:
            skey = (company.name, d, sd, s)
            section_obj = section_map.get(skey)
            if not section_obj:
                section_obj = Section(name=s, sub_department=subdept_obj, employee_count=0, manager=None)
                sections_to_create.append(section_obj)
                section_map[skey] = section_obj

        if d and sd and s and ss:
            sskey = (company.name, d, sd, s, ss)
            if sskey not in subsection_map:
                subsections_to_create.append(SubSection(name=ss, section=section_obj, employee_count=0, manager=None))
                subsection_map[sskey] = True  # placeholder

    if dry_run:
        return {
            "status": "ok",
            "to_create": {
                "departments": len(depts_to_create),
                "sub_departments": len(subdepts_to_create),
                "sections": len(sections_to_create),
                "sub_sections": len(subsections_to_create),
            }
        }

    # Commit in dependency order
    with transaction.atomic():
        if depts_to_create:
            Department.objects.bulk_create(depts_to_create, ignore_conflicts=True)
            dept_map = _fetch_departments(company_names, dept_keys)  # refresh

        if subdepts_to_create:
            for sd in subdepts_to_create:
                sd.department = dept_map[(sd.department.company.name, sd.department.name)]
            SubDepartment.objects.bulk_create(subdepts_to_create, ignore_conflicts=True)
            subdept_map = _fetch_subdepartments(company_names, subdept_keys)  # refresh

        if sections_to_create:
            for s in sections_to_create:
                s.sub_department = subdept_map[(s.sub_department.department.company.name, s.sub_department.department.name, s.sub_department.name)]
            Section.objects.bulk_create(sections_to_create, ignore_conflicts=True)
            section_map = _fetch_sections(company_names, section_keys)

        if subsections_to_create:
            for ss in subsections_to_create:
                ss.section = section_map[(ss.section.sub_department.department.company.name, ss.section.sub_department.department.name, ss.section.sub_department.name, ss.section.name)]
            SubSection.objects.bulk_create(subsections_to_create, ignore_conflicts=True)

    return {
        "status": "imported",
        "created": {
            "departments": len(depts_to_create),
            "sub_departments": len(subdepts_to_create),
            "sections": len(sections_to_create),
            "sub_sections": len(subsections_to_create),
        }
    }


# ---------- Internal helpers ----------

def _normalize_and_select_headers(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Map messy headers from the sheet to canonical keys:
      company, department, sub_department, section, sub_section
    """
    def pick(d, *keys):
        for k in keys:
            if k in d and d[k] not in [None, ""]:
                return str(d[k]).strip()
        return None

    cleaned = []
    for i, r in enumerate(rows, start=1):
        cleaned.append({
            "__row": i,
            "company":       pick(r, "company", "Company", "company_name", "Company Name"), 
            "department":    pick(r, "department", "Department", "Department' Name", "Department Name"),
            "sub_department":pick(r, "sub_department", "Sub Department", "SubDepartment", "Sub Department'name"),
            "section":       pick(r, "section", "Section", "Section'name"),
            "sub_section":   pick(r, "sub_section", "Sub Section", "SubSection"),
        })
    return cleaned


def _basic_validate(cleaned: List[Dict[str, Any]]):
    errors = []
    company_names = set()
    for row in cleaned:
        if not row["company"]:
            errors.append({"row": row["__row"], "errors": {"company": ["This field is required."]}})
        if not any([row["department"], row["sub_department"], row["section"], row["sub_section"]]):
            errors.append({"row": row["__row"], "errors": {"path": ["Provide at least one level below company."]}})
        if row["company"]:
            company_names.add(row["company"])
    return errors, company_names


def _collect_unique_keys(cleaned: List[Dict[str, Any]]):
    dept_keys, subdept_keys, section_keys, subsection_keys = set(), set(), set(), set()
    for r in cleaned:
        c = r["company"]; d = r["department"]; sd = r["sub_department"]; s = r["section"]; ss = r["sub_section"]
        if d:                   dept_keys.add((c, d))
        if d and sd:           subdept_keys.add((c, d, sd))
        if d and sd and s:     section_keys.add((c, d, sd, s))
        if d and sd and s and ss: subsection_keys.add((c, d, sd, s, ss))
    return dept_keys, subdept_keys, section_keys, subsection_keys 


def _fetch_departments(company_names, dept_keys) -> Dict[Any, Department]:
    if not dept_keys:
        return {}
    qs: QuerySet[Department] = Department.objects.filter(
        company__name__in=company_names,
        name__in=[k[1] for k in dept_keys]
    ).select_related("company")
    return {(d.company.name, d.name): d for d in qs}


def _fetch_subdepartments(company_names, subdept_keys) -> Dict[Any, SubDepartment]:
    if not subdept_keys:
        return {}
    qs: QuerySet[SubDepartment] = SubDepartment.objects.filter(
        department__company__name__in=company_names,
        name__in=[k[2] for k in subdept_keys]
    ).select_related("department", "department__company")
    return {(x.department.company.name, x.department.name, x.name): x for x in qs}


def _fetch_sections(company_names, section_keys) -> Dict[Any, Section]:
    if not section_keys:
        return {}
    qs: QuerySet[Section] = Section.objects.filter(
        sub_department__department__company__name__in=company_names,
        name__in=[k[3] for k in section_keys]
    ).select_related("sub_department", "sub_department__department", "sub_department__department__company")
    return {(x.sub_department.department.company.name, x.sub_department.department.name, x.sub_department.name, x.name): x for x in qs}


def _fetch_subsections(company_names, subsection_keys) -> Dict[Any, SubSection]:
    if not subsection_keys:
        return {}
    qs: QuerySet[SubSection] = SubSection.objects.filter(
        section__sub_department__department__company__name__in=company_names,
        name__in=[k[4] for k in subsection_keys]
    ).select_related("section", "section__sub_department", "section__sub_department__department", "section__sub_department__department__company")
    return {(x.section.sub_department.department.company.name, x.section.sub_department.department.name, x.section.sub_department.name, x.section.name, x.name): x for x in qs}
