# app/services/placement_levels_importer.py
from __future__ import annotations

from pathlib import Path
from io import TextIOWrapper
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime, date
import csv

try:
    import openpyxl
except ImportError:
    openpyxl = None

from django.db import transaction
from django.utils import timezone
from django.db.models import Q

from evaluation_app.models import (
    Employee, EmployeePlacement,
    Company, Department, SubDepartment, Section, SubSection,
)

# ---------------------- I/O (CSV/XLSX/JSON) ----------------------

def parse_levels_rows(request) -> List[Dict[str, Any]]:
    """
    Accept either:
      - multipart/form-data with key 'file' (CSV or XLSX)
      - application/json body = array of rows

    Expected headers (case-insensitive, flexible):
      Emp Code | Department | Sub Department | Section | Sub Section
    """
    if "file" in request.FILES:
        f = request.FILES["file"]
        suffix = Path(f.name).suffix.lower()

        if suffix == ".csv":
            text = TextIOWrapper(f.file, encoding="utf-8", newline="")
            reader = csv.DictReader(text)
            rows = [row for row in reader if any(v not in [None, ""] for v in row.values())]
            if not rows:
                raise ValueError("CSV appears empty.")
            return rows

        if suffix in {".xlsx", ".xls"}:
            if openpyxl is None:
                raise ValueError("XLSX import requires 'openpyxl'. Install it or upload CSV.")
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active

            try:
                header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                raise ValueError("XLSX appears empty (no header row).")

            headers = [(str(c.value).strip() if c.value is not None else "") for c in header_cells]

            rows: List[Dict[str, Any]] = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r is None or all(v in (None, "") for v in r):
                    continue
                rows.append({headers[i]: (r[i] if i < len(headers) else None) for i in range(len(headers))})

            if not rows:
                raise ValueError("XLSX sheet appears empty.")
            return rows

        raise ValueError("Unsupported file type. Please upload CSV or XLSX.")

    data = request.data
    if not isinstance(data, list):
        raise ValueError('Expected a JSON array or upload a file as "file".')
    return data


# ---------------------- Core Import Logic ----------------------

def import_levels(
    rows: List[Dict[str, Any]],
    *,
    dry_run: bool = False,
    effective: Optional[str] = None,   # "YYYY-MM-DD" → assigned_at
) -> Dict[str, Any]:
    """
    Bulk update (or create) EmployeePlacement levels by Emp Code and unit names.
    One placement per employee is assumed (your unique constraint). If missing, it will be created.
    """
    cleaned = _clean_rows(rows)
    bad = _validate_rows(cleaned)
    if bad:
        return {"status": "invalid", "errors": bad}

    # Resolve effective date once
    assigned_at = None
    if effective:
        try:
            dt = datetime.strptime(effective, "%Y-%m-%d")
            assigned_at = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        except Exception:
            return {"status": "invalid", "errors": [{"row": None, "errors": {"effective": ["Use YYYY-MM-DD"]}}]}

    # Prefetch employees by code (may be ambiguous across companies; we’ll disambiguate by path later)
    codes = [r["employee_code"] for r in cleaned]
    employees = (
        Employee.objects
        .filter(employee_code__in=codes)
        .select_related("company", "user")
    )

    by_code: Dict[str, List[Employee]] = {}
    for e in employees:
        by_code.setdefault(e.employee_code, []).append(e)

    # Caches to avoid re-querying by name for the same company/parent
    dept_cache: Dict[Tuple[str, str], Optional[Department]] = {}        # (company_id, norm_name) -> Department
    sdep_cache: Dict[Tuple[str, str], Optional[SubDepartment]] = {}     # (department_id, norm_name) -> SubDepartment
    sec_cache: Dict[Tuple[str, str], Optional[Section]] = {}            # (sub_department_id, norm_name) -> Section
    ssec_cache: Dict[Tuple[str, str], Optional[SubSection]] = {}        # (section_id, norm_name) -> SubSection

    created = updated = unchanged = 0
    errors: List[Dict[str, Any]] = []
    to_create: List[EmployeePlacement] = []
    to_update: List[EmployeePlacement] = []

    # Helper resolvers (cached, case-insensitive)
    def norm(s: Optional[str]) -> Optional[str]:
        return s.strip().lower() if isinstance(s, str) else None

    def get_department(company_id: str, name: Optional[str]) -> Optional[Department]:
        if not name:
            return None
        key = (company_id, norm(name))
        if key in dept_cache:
            return dept_cache[key]
        obj = Department.objects.filter(company_id=company_id, name__iexact=name).first()
        dept_cache[key] = obj
        return obj

    def get_subdept(dept_id: str, name: Optional[str]) -> Optional[SubDepartment]:
        if not name:
            return None
        key = (dept_id, norm(name))
        if key in sdep_cache:
            return sdep_cache[key]
        obj = SubDepartment.objects.filter(department_id=dept_id, name__iexact=name).first()
        sdep_cache[key] = obj
        return obj

    def get_section(sdep_id: str, name: Optional[str]) -> Optional[Section]:
        if not name:
            return None
        key = (sdep_id, norm(name))
        if key in sec_cache:
            return sec_cache[key]
        obj = Section.objects.filter(sub_department_id=sdep_id, name__iexact=name).first()
        sec_cache[key] = obj
        return obj

    def get_subsection(sec_id: str, name: Optional[str]) -> Optional[SubSection]:
        if not name:
            return None
        key = (sec_id, norm(name))
        if key in ssec_cache:
            return ssec_cache[key]
        obj = SubSection.objects.filter(section_id=sec_id, name__iexact=name).first()
        ssec_cache[key] = obj
        return obj

    # Prefetch existing placements for all involved employees (there should be at most one by constraint)
    # We cannot use in_bulk(field_name="employee_id") since it's not a unique field in the model definition.
    placements_qs = (
        EmployeePlacement.objects
        .filter(employee__employee_code__in=codes)
        .select_related("company", "department", "sub_department", "section", "sub_section")
        .order_by("employee_id", "-assigned_at", "-placement_id")
    )
    placement_by_emp: Dict[str, EmployeePlacement] = {}
    for p in placements_qs:
        placement_by_emp.setdefault(str(p.employee_id), p)

    for r in cleaned:
        row_no = r["__row"]
        code = r["employee_code"]
        dname, sdname, sname, ssname = r["department"], r["sub_department"], r["section"], r["sub_section"]

        candidates = by_code.get(code, [])
        if not candidates:
            errors.append({"row": row_no, "errors": {"Emp Code": [f'Employee "{code}" not found.']}})
            continue

        # Choose the employee whose COMPANY actually contains the provided path
        matched_emp: Optional[Employee] = None
        matched_company: Optional[Company] = None
        dep = sdep = sec = ssec = None

        for cand in candidates:
            if not cand.company_id:
                # No company on employee → cannot resolve this candidate
                continue

            # Resolve chain under this company
            dep = get_department(str(cand.company_id), dname) if dname else None
            if dname and not dep:
                continue

            sdep = get_subdept(str(dep.pk), sdname) if (dep and sdname) else None
            if sdname and not sdep:
                continue

            sec = get_section(str(sdep.pk), sname) if (sdep and sname) else None
            if sname and not sec:
                continue

            ssec = get_subsection(str(sec.pk), ssname) if (sec and ssname) else None
            if ssname and not ssec:
                continue

            if matched_emp is None:
                matched_emp = cand
                matched_company = cand.company
            else:
                errors.append({"row": row_no, "errors": {"Emp Code": [f'Code "{code}" is used in multiple companies and the path matches more than one.']}})
                matched_emp = None
                break

        if matched_emp is None:
            errors.append({"row": row_no, "errors": {"path": ["Path not found under employee’s company (or employee has no company)."]}})
            continue

        # Get/create that employee's single placement
        placement = placement_by_emp.get(str(matched_emp.pk))

        # Compute target signature for idempotency
        target_sig = (
            matched_company.pk if matched_company else None,
            dep.pk if dep else None,
            sdep.pk if sdep else None,
            sec.pk if sec else None,
            ssec.pk if ssec else None,
        )

        if placement:
            current_sig = (
                placement.company_id,
                placement.department_id,
                placement.sub_department_id,
                placement.section_id,
                placement.sub_section_id,
            )
            if current_sig == target_sig:
                unchanged += 1
                continue

            # Update in memory
            placement.company        = matched_company
            placement.department     = dep if dname else None
            placement.sub_department = sdep if sdname else None
            placement.section        = sec if sname else None
            placement.sub_section    = ssec if ssname else None
            if assigned_at is not None:
                placement.assigned_at = assigned_at
            to_update.append(placement)
            updated += 1
        else:
            # Create in memory
            newp = EmployeePlacement(
                employee        = matched_emp,
                company         = matched_company,
                department      = dep if dname else None,
                sub_department  = sdep if sdname else None,
                section         = sec if sname else None,
                sub_section     = ssec if ssname else None,
                assigned_at     = assigned_at if assigned_at is not None else timezone.now(),
            )
            to_create.append(newp)
            created += 1

    if errors:
        return {"status": "invalid", "errors": errors}

    if dry_run:
        return {
            "status": "ok",
            "validated": len(cleaned),
            "to_create": len(to_create),
            "to_update": len(to_update),
            "unchanged": unchanged,
        }

    with transaction.atomic():
        if to_create:
            EmployeePlacement.objects.bulk_create(to_create)
        if to_update:
            EmployeePlacement.objects.bulk_update(
                to_update,
                ["company", "department", "sub_department", "section", "sub_section", "assigned_at"],
            )

    return {
        "status": "imported",
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
    }


# ---------------------- Helpers ----------------------

def _clean_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Normalize headers & cell values. Keep only relevant columns."""
    def pick(d, *keys):
        for k in keys:
            if k in d and d[k] not in (None, ""):
                return d[k]
        return None

    cleaned: List[Dict[str, Any]] = []
    for i, r in enumerate(rows, start=1):
        cleaned.append({
            "__row": i,
            "employee_code": _s(pick(r, "Emp Code", "EmpCode", "employee_code", "Employee Code")),
            "department": _s(pick(r, "Department", "Dept")),
            "sub_department": _s(pick(r, "Sub Department", "Sub_Department", "SubDepartment")),
            "section": _s(pick(r, "Section")),
            "sub_section": _s(pick(r, "Sub Section", "Sub_Section", "SubSection")),
        })
    return cleaned

def _validate_rows(cleaned: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    errs: List[Dict[str, Any]] = []
    for row in cleaned:
        local: Dict[str, List[str]] = {}
        if not row["employee_code"]:
            local.setdefault("Emp Code", []).append("This field is required.")
        if not any([row["department"], row["sub_department"], row["section"], row["sub_section"]]):
            local.setdefault("path", []).append("Provide at least Department or deeper.")
        if row["sub_department"] and not row["department"]:
            local.setdefault("Department", []).append("Required when Sub Department is provided.")
        if row["section"] and not row["sub_department"]:
            local.setdefault("Sub Department", []).append("Required when Section is provided.")
        if row["sub_section"] and not row["section"]:
            local.setdefault("Section", []).append("Required when Sub Section is provided.")
        if local:
            errs.append({"row": row["__row"], "errors": local})
    return errs

def _s(v) -> Optional[str]:
    if v is None:
        return None
    vs = str(v).strip()
    return vs or None
