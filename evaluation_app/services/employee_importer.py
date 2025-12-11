# app/services/employee_importer.py
from __future__ import annotations
from pathlib import Path
from io import TextIOWrapper
import csv, re
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime, date
from difflib import get_close_matches

try:
    import openpyxl
except ImportError:
    openpyxl = None

from django.db import transaction
from django.contrib.auth import get_user_model
from evaluation_app.serializers.employee_serilized import EmployeeSerializer  # local import to avoid cycles

from evaluation_app.models import (
    Company, Department, Employee,
    # enums below must exist in your project
    ManagerialLevel, EmpStatus, JobType, BranchType  # adjust import
)
from accounts.models import User, Gender  # adjust import

User = get_user_model()

# ------------------ Public API ------------------

def parse_employee_rows(request) -> List[Dict[str, Any]]:
    """Return list[dict] from JSON array OR multipart CSV/XLSX under 'file' key."""
    if "file" in request.FILES:
        f = request.FILES["file"]
        suffix = Path(f.name).suffix.lower()
        if suffix == ".csv":
            text = TextIOWrapper(f.file, encoding="utf-8", newline="")
            reader = csv.DictReader(text)
            rows = list(reader)
            if not rows:
                raise ValueError("CSV appears empty.")
            return rows
        if suffix in {".xlsx", ".xls"}:
            if openpyxl is None:
                raise ValueError("XLSX import requires 'openpyxl'. Install it or upload CSV.")
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in r):
                    continue
                rows.append({headers[i]: r[i] for i in range(len(headers))})
            if not rows:
                raise ValueError("XLSX sheet appears empty.")
            return rows
        raise ValueError("Unsupported file type. Upload CSV or XLSX.")
    # JSON
    data = request.data
    if not isinstance(data, list):
        raise ValueError('Expected a JSON array or upload a file as "file".')
    return data


def import_employees(rows: List[Dict[str, Any]], *, dry_run: bool = False,
                     update_existing: bool = True, upsert_by_code: bool = True) -> Dict[str, Any]:
    """
    Bulk import employees.
    Upsert key = email; optionally also match by (company, employee_code) if upsert_by_code=True.
    Companies & Departments must already exist.
    """
    cleaned, companies_needed, dept_keys = _clean_and_normalize(rows)

    # Validate presence of required cols per row
    errors = _validate_rows(cleaned)
    if errors:
        return {"status": "invalid", "errors": errors}

    # Prefetch companies
    companies = {c.name: c for c in Company.objects.filter(name__in=companies_needed)}
    missing_companies = [n for n in companies_needed if n not in companies]
    if missing_companies:
        errs = []
        for row in cleaned:
            if row["company_name"] in missing_companies:
                errs.append({"row": row["__row"], "errors": {"Company Name": [f'Company "{row["company_name"]}" does not exist. Import companies first.']}})
        return {"status": "invalid", "errors": errs}

    # Prefetch departments (by company+name)
    dept_map: Dict[Tuple[str, str], Department] = {}
    if dept_keys:
        existing_depts = Department.objects.filter(
            company__name__in=[k[0] for k in dept_keys],
            name__in=[k[1] for k in dept_keys],
        ).select_related("company")
        dept_map = {(d.company.name, d.name): d for d in existing_depts}

    # Prefetch users by email
    emails = {r["email"] for r in cleaned if r["email"]}
    users_by_email = {u.email: u for u in User.objects.filter(email__in=emails)}

    # Prefetch employees by (company, employee_code) if desired
    empcode_keys = {(row["company_name"], row["employee_code"]) for row in cleaned
                    if upsert_by_code and row.get("employee_code")}
    employees_by_code: Dict[Tuple[str, str], Employee] = {}
    if empcode_keys:
        employees = Employee.objects.filter(
            company__name__in=[k[0] for k in empcode_keys],
            employee_code__in=[k[1] for k in empcode_keys],
        ).select_related("company", "user")
        employees_by_code = {(e.company.name if e.company else "", e.employee_code): e for e in employees}

    # Prepare create/update plans
    to_create: List[Tuple[int, Dict[str, Any]]] = []
    to_update: List[Tuple[int, Employee, Dict[str, Any]]] = []

    # Also keep track of usernames to avoid duplicates when creating Users
    taken_usernames = set(User.objects.filter(username__in=[_username_from_email(e) for e in emails]).values_list("username", flat=True))

    row_errors = []
    for row in cleaned:
        row_no = row["__row"]
        company = companies[row["company_name"]]
        # Department resolution (optional)
        dept_obj = None
        if row["department"]:
            dkey = (company.name, row["department"])
            dept_obj = dept_map.get(dkey)
            if not dept_obj:
                row_errors.append({"row": row["__row"], "errors": {"Department": [f'Department "{row["department"]}" not found in company "{company.name}".']}})
                continue

        # Upsert resolution
        user = users_by_email.get(row["email"])
        employee = None
        if user and hasattr(user, "employee_profile"):
            employee = user.employee_profile
        elif upsert_by_code and row.get("employee_code"):
            employee = employees_by_code.get((company.name, row["employee_code"]))

        emp_payload, user_payload = _build_payloads(row, company, dept_obj, taken_usernames)

       # if dry_run: (no more needed)
       #     continue  # Only validating structure; real field validity is enforced in serializer at commit-time 

        if employee:
            # UPDATE: patch employee + user, NO password
            data = dict(emp_payload)
            data.pop("user_data", None)      # don't use create-only field on update
            data["user"] = user_payload      # EmployeeSerializer.update() expects 'user'
            to_update.append((row_no, employee, data)) 
        elif user:
            # CREATE new Employee bound to EXISTING user (no password change)
           data = dict(emp_payload)
           data.pop("user_data", None)      # we aren't creating a user
           data["user_id"] = str(user.pk)   # attach by ID
           to_create.append((row_no, data))
        else:
           # CREATE brand-new User + Employee → set default password
           data = dict(emp_payload)
           data.setdefault("user_data", user_payload)
           data["user_data"]["password"] = "defaultpassword123"  # default password for new users
           to_create.append((row_no,data))

    if row_errors:
        return {"status": "invalid", "errors": row_errors}
    
    # ---------- PRE-VALIDATION (works for dry_run and commit) ----------
    validation_errors = []
    # Validate creates
    for row_no, data in to_create:
        ser = EmployeeSerializer(data=data)
        if not ser.is_valid():
            validation_errors.append({"row": row_no, "errors": _pretty_errors(ser.errors, data)})


    # Validate updates
    for row_no, instance, data in to_update:     
        ser = EmployeeSerializer(instance, data=data, partial=True)
        if not ser.is_valid():
            validation_errors.append({"row": row_no, "errors": _pretty_errors(ser.errors, data)})   

    if validation_errors:
        return{
            "status": "invalid",
            "errors": validation_errors,
            "accepted_values": _accepted_values_hint(),
        }     
       
    if dry_run:
        return {"status": "imported", 
                "validated_count": len(cleaned),
                "to_create": len(to_create),
                "to_update": len(to_update),}
    
    # ---------- COMMIT ----------
    created = updated = 0

    # Write all in a transaction
    with transaction.atomic():
        # Create
        for _, data in to_create:
            ser = EmployeeSerializer(data=data)
            ser.is_valid(raise_exception=True) # should always be valid here
            ser.save()
            created += 1

        # Update
        for _, instance, data in to_update:
            ser = EmployeeSerializer(instance, data=data, partial=True)
            ser.is_valid(raise_exception=True)
            ser.save()
            updated += 1

    return {"status": "imported", "created": created, "updated": updated}


# ------------------ Helpers ------------------

def _username_from_email(email: str) -> str:
    base = (email or "").split("@")[0][:150] or "user"
    # basic cleanup
    base = re.sub(r"[^a-zA-Z0-9_.-]+", "", base)
    return base or "user"

def _unique_username(suggested: str, taken: set) -> str:
    username = suggested
    n = 1
    while username in taken or User.objects.filter(username=username).exists():
        n += 1
        username = f"{suggested}-{n}"
    taken.add(username)
    return username

def _parse_date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    # try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # last resort: let serializer fail with a clear message
    return s

def _norm_choice(value, choices, aliases=None):
    """
    Accept either stored values, labels, or loose variants (spaces/hyphens ignored).
    `aliases` lets you add custom synonyms like "full time" -> FULL_TIME.
    """
    if value is None or value == "":
        return None
    key = _norm_key(value)
    table = {}
    # stored values
    for v, lbl in choices:
        table[_norm_key(v)] = v
        table[_norm_key(lbl)] = v
    # custom aliases
    if aliases:
        for alias, real in aliases.items():
            table[_norm_key(alias)] = real
    return table.get(key, value)  # fall through: let serializer complain if still unknown

JOBTYPE_ALIASES = {
    "full time": JobType.FULL_TIME,
    "fulltime": JobType.FULL_TIME,
    "part time": JobType.PART_TIME,
    "parttime": JobType.PART_TIME,
    "full time remote": JobType.FULL_TIME_REMOTE,
    "remote full time": JobType.FULL_TIME_REMOTE,
    "full remote": JobType.FULL_TIME_REMOTE,
    "part time remote": JobType.PART_TIME_REMOTE,
    "remote part time": JobType.PART_TIME_REMOTE,
}

STATUS_ALIASES = {
    "defaultactive": EmpStatus.DEFAULT,   # in case sheet says "Default Active"
    "default active": EmpStatus.DEFAULT,
    "active": EmpStatus.ACTIVE,
    "inactive": EmpStatus.INACTIVE,
}

MANAGERIAL_ALIASES = {
    "ic": ManagerialLevel.IC,
    "individual contributor": ManagerialLevel.IC,
    "supervisory": ManagerialLevel.SUPERVISORY,
    "middle": ManagerialLevel.MIDDLE,
    "middle management": ManagerialLevel.MIDDLE,
    "executive": ManagerialLevel.EXECUTIVE,
    "executive management": ManagerialLevel.EXECUTIVE,
}

BRANCH_ALIASES = {
    "office": BranchType.OFFICE,
    "store": BranchType.STORE,
}

GENDER_ALIASES = {
    "m": Gender.MALE,
    "male": Gender.MALE,
    "f": Gender.FEMALE,
    "female": Gender.FEMALE,
}

def _clean_and_normalize(rows: List[Dict[str, Any]]):
    def pick(d, *keys):
        for k in keys:
            if k in d and d[k] not in [None, ""]:
                return str(d[k]).strip() if not isinstance(d[k], (date, datetime)) else d[k]
        return None

    cleaned = []
    companies_needed = set()
    dept_keys = set()

    for i, r in enumerate(rows, start=1):
        row = {
            "__row": i,
            "employee_code": pick(r, "Emp Code", "EmpCode", "employee_code"),
            "name": pick(r, "Name", "Full Name", "Employee Name"),
            "email": pick(r, "Email", "E-mail"),
            "country_code": pick(r, "CountryCode", "Country Code"),
            "phone": pick(r, "Phone Number", "Phone", "Mobile"),
            "department": pick(r, "Department", "Dept"),
            "position": pick(r, "Position"),
            "role": pick(r, "Role"),
            "managerial_level": pick(r, "Managerial Level", "ManagerialLevel"),
            "company_name": pick(r, "Company Name", "Company"),
            "status": pick(r, "Status"),
            "join_date": pick(r, "Join Date", "JoinDate", "Start Date"),
            "location": pick(r, "Location"),
            "branch": pick(r, "Branch"),
            "job_type": pick(r, "Job Type", "JobType"),
            "gender": pick(r, "Gender", "Sex"),
            # ignored: Profile Image, Direct Manager\Supervisor, Head of Dep
        }
        # Post-process
        row["join_date"] = _parse_date(row["join_date"])
        # normalize choices
        row["role"]             = _norm_choice(row["role"], User._meta.get_field("role").choices)  # labels/values both OK
        row["managerial_level"] = _norm_choice(row["managerial_level"], ManagerialLevel.choices, MANAGERIAL_ALIASES)
        row["status"]           = _norm_choice(row["status"], EmpStatus.choices, STATUS_ALIASES)
        row["job_type"]         = _norm_choice(row["job_type"], JobType.choices, JOBTYPE_ALIASES)
        row["branch"]           = _norm_choice(row["branch"], BranchType.choices, BRANCH_ALIASES)
        row["gender"]           = _norm_choice(row["gender"], Gender.choices, GENDER_ALIASES)
        cleaned.append(row)

        if row["company_name"]:
            companies_needed.add(row["company_name"])
        if row["company_name"] and row["department"]:
            dept_keys.add((row["company_name"], row["department"]))

    return cleaned, companies_needed, dept_keys


def _validate_rows(cleaned: List[Dict[str, Any]]):
    errs = []
    for row in cleaned:
        local = {}
        if not row["name"]:
            local.setdefault("Name", []).append("This field is required.")
        if not row["email"]:
            local.setdefault("Email", []).append("This field is required.")
        if not row["company_name"]:
            local.setdefault("Company Name", []).append("This field is required.")
        if not row["managerial_level"]:
            local.setdefault("Managerial Level", []).append("This field is required.")
        if not row["status"]:
            local.setdefault("Status", []).append("This field is required.")
        if not row["join_date"]:
            local.setdefault("Join Date", []).append("This field is required.")
        if local:
            errs.append({"row": row["__row"], "errors": local})
    return errs


def _build_payloads(row, company: Company, dept: Optional[Department], taken_usernames: set):
    # User payload
    username = _unique_username(_username_from_email(row["email"]), taken_usernames)
    user_payload = {
        "username": username,
        "name": row["name"],
        "email": row["email"],
        "phone": row.get("phone") or "",
        "country_code": row.get("country_code") or "",
        "role": row.get("role"),
        "position": row.get("position") or "",
        # Optional: include gender if your serializer exposes it
        "gender": row.get("gender"),
        # No password provided: creates an unusable password; set later via invite
    }
    # Employee payload
    emp_payload = {
        "company_id": str(company.pk),
        "department_id": str(dept.pk) if dept else None,
        "employee_code": row.get("employee_code") or "",
        "managerial_level": row["managerial_level"],
        "status": row["status"],
        "join_date": row["join_date"],  # '%Y-%m-%d'
        "job_type": row.get("job_type"),
        "location": row.get("location") or "",
        "branch": row.get("branch"),
        "user_data": user_payload,
    }
    return emp_payload, user_payload

def _norm_key(s: str) -> str:
    # normalize strings for matching: lowercase & strip all non-alphanumerics
    return re.sub(r'[^a-z0-9]+', '', str(s).lower())





def _accepted_values_hint():
    return {
        "managerial_level": [lbl for _, lbl in ManagerialLevel.choices],
        "status": [lbl for _, lbl in EmpStatus.choices],
        "job_type": [lbl for _, lbl in JobType.choices],
        "branch": [lbl for _, lbl in BranchType.choices],
        "gender": [lbl for _, lbl in Gender.choices],
        "role": [lbl for _, lbl in User._meta.get_field("role").choices],
    }

def _closest(label_list, raw):
    if not raw:
        return None
    m = get_close_matches(str(raw), label_list, n=1, cutoff=0.6)
    return m[0] if m else None

def _pretty_errors(errors_dict, original_row_payload):
    """
    Optionally enrich DRF errors like 'is not a valid choice.' with suggestions
    based on your choices labels.
    """
    out = {}
    for field, msgs in errors_dict.items():
        if isinstance(msgs, (list, tuple)):
            text = [str(x) for x in msgs]
        else:
            text = [str(msgs)]

        # add "did you mean" for common choice fields
        if field in ("managerial_level", "status", "job_type", "branch", "gender", "role"):
            labels = _accepted_values_hint()[field]
            raw = original_row_payload.get(field)
            hint = _closest(labels, raw)
            if hint and any("valid choice" in m.lower() for m in text):
                text.append(f"Did you mean “{hint}”?")

        out[field] = text
    return out
