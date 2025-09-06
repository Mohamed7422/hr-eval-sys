from evaluation_app.serializers.org_serializers import(
    CompanySerializer, DepartmentSerializer, SubDepartmentSerializer, SectionSerializer, SubSectionSerializer, EmployeePlacementSerializer
)
from evaluation_app.permissions import IsAdmin, IsHR, IsHOD, IsLineManager, IsSelfOrAdminHR, ReadOnlyOrAdminHR,IsAdminOrHR
from evaluation_app.models import Employee, Company, Department, SubDepartment, Section, SubSection, EmployeePlacement, CompanySize
from rest_framework import viewsets, filters, permissions, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
from pathlib import Path
import csv
from io import TextIOWrapper
from django.shortcuts import get_object_or_404
try:
    import openpyxl
except ImportError:
    openpyxl = None 
    
from django_filters.rest_framework import DjangoFilterBackend
from evaluation_app.services.hierarchy_importer import parse_hierarchy_rows, import_hierarchy_paths



class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all().order_by("name")
    serializer_class = CompanySerializer
    permission_classes = [ReadOnlyOrAdminHR] # read-only for authenticated users, full access for Admin/HR
    #print("CompanyViewSet permissions:", permission_classes)
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "industry"]
     # … list / create / retrieve / update / destroy are inherited …

    @action(
      detail=False, #collection-level
      methods=["post"],
      url_path="create",  #  → /companies/create/
      permission_classes=[permissions.IsAuthenticated,IsAdminOrHR]
    )
    def create_company(self, request, *args, **kwargs):
        """
        Custom action to create a new company.
        Only Admin or HR can create a company.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    #------------------ Importing At Once Section ------------------

    
    @action(
            detail=False,
            methods=["post"],
            url_path="import",  # → /companies/import/
            permission_classes=[permissions.IsAuthenticated, IsAdminOrHR]
    )
    def import_companies(self, request, *args, **kwargs):
        """
        Bulk import companies from JSON array or uploaded CSV/XLSX file.

        - JSON: POST array of objects
        - File: multipart/form-data with 'file'
        Query params:
          - dry_run=true     : validate only
          - update_existing=true : upsert by 'name' (change key below if needed)
        """
        dry_run = request.query_params.get("dry_run") == "true"
        update_existing = request.query_params.get("update_existing") == "true"
        upsert_on = "name"

        #1) pars rows from JSON or file
        try:
            rows = self._parse_payload_to_rows(request)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # 2) Normalize values (e.g., size) + keep only model fields
        cleaned = []
        for r in rows:
            item = {
                 "name": r.get("name") or r.get("Name"),
                 "industry": r.get("industry") or r.get("Industry"),
                 "size": self._normalize_size(r.get("size") or r.get("Size")),
                 "address": r.get("address") or r.get("Address"),
                #NOTE: 'Description' is ignored because the model has no 'description' field.
            }
            cleaned.append(item) 

        # 3) Validate row-by-row to report per-row errors
        errors = []
        valid_instances = []
        
        for idx, item in enumerate(cleaned, start=1):
            ser = self.get_serializer(data=item)
            if ser.is_valid():
                valid_instances.append(ser.validated_data)
            else:
                errors.append({"row": idx, "errors": ser.errors})

        if errors:
            return Response(
                {"status": "invalid", "errors": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if dry_run:
            return Response({"status": "ok", "validated_count_existing": len(valid_instances)})

        # 4) Write to DB (all-or-nothing)
        created = []
        updated = []
        with transaction.atomic():
            if update_existing:
                for data in valid_instances:
                    # upsert by 'upsert_on'
                    lookup = {upsert_on: data[upsert_on]}
                    obj, was_created = Company.objects.update_or_create(
                        defaults=data, **lookup
                    )
                    (created if was_created else updated).append(obj)
            else:
                objs = [Company(**data) for data in valid_instances]
                Company.objects.bulk_create(objs)
                created.extend(objs)

        # 5) Return summary
        created_data = self.get_serializer(created, many=True).data
        updated_data = self.get_serializer(updated, many=True).data
        return Response(
            {
                "status": "imported",
                "created": len(created),
                "updated": len(updated),
                "items_created": created_data,
                "items_updated": updated_data,
            },
            status=status.HTTP_201_CREATED,
        )
    #-------------------------------------------
    def _parse_payload_to_rows(self, request):
        """
        Returns a list[dict] from either:
        - JSON array body, or
        - multipart file 'file' with CSV/XLSX.
        """   
        if "file" in request.FILES: 
            f = request.FILES["file"]
            suffix = Path(f.name).suffix.lower()
            if suffix == ".csv":
            # Expect UTF-8 CSV with header row: name,Industry,size,Description,Address
               text = TextIOWrapper(f.file, encoding="utf-8", newline="")
               reader = csv.DictReader(text)
               rows = list(reader)
               if not rows:
                   raise ValueError("Uploaded CSV file is empty")
               return rows  
            
            if suffix in {".xlsx", ".xls"}:
               if openpyxl is None:
                  raise ValueError("XLSX import requires openpyxl. Install it or upload CSV.")
               wb = openpyxl.load_workbook(f, read_only=True)
               ws = wb.active
               headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
               rows = []
               for row in ws.iter_rows(min_row=2, values_only=True):
                   if all(v is None for v in row):
                        continue
                   rows.append({headers[i]: row[i] for i in range(len(headers))})
               if not rows:
                    raise ValueError("XLSX sheet appears to be empty.")
               return rows

            raise ValueError("Unsupported file type. Upload CSV or XLSX.")
        else:
            data = request.data
            if not isinstance(data, list):
                raise ValueError('Expected a JSON array or upload a file as "file".')
            return data

    def _normalize_size(self, raw):
        """
        Accept either the stored value or the display label for CompanySize.
        Falls back to raw so the serializer can raise a proper choice error if unknown.
        """
        if raw is None:
            return raw
        s = str(raw).strip()
        # Map both values and labels (case-insensitive) to stored value
        value_by_value = {str(v).lower(): v for v, _ in CompanySize.choices}
        value_by_label = {str(lbl).lower(): v for v, lbl in CompanySize.choices}
        key = s.lower()
        return value_by_value.get(key) or value_by_label.get(key) or s
    # ──────────────────────────────────────────────────────────────────────────────
    # Bulk import of full hierarchy paths
    # ──────────────────────────────────────────────────────────────────────────────
    @action(
        detail=False,
        methods=["post"],
        url_path="import-hierarchy",
        permission_classes=[permissions.IsAuthenticated, IsAdminOrHR],
    )
    def import_hierarchy(self, request, *args, **kwargs):
        dry_run = request.query_params.get("dry_run") == "true"
        try:
            rows = parse_hierarchy_rows(request)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        result = import_hierarchy_paths(rows, dry_run=dry_run)

        if result.get("status") == "invalid":
            return Response(result, status=status.HTTP_400_BAD_REQUEST)

        http_status = status.HTTP_200_OK if result["status"] == "ok" else status.HTTP_201_CREATED
        return Response(result, status=http_status)


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.select_related("company", "manager")
    serializer_class = DepartmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = {
        "company": ["exact"],
        "company_id": ["exact"],
    }
    search_fields = ["name"]

    @action(
        detail=False,  # collection-level
        methods=["post"],
        url_path="create",  # → /departments/create/
        permission_classes=[permissions.IsAuthenticated, IsAdminOrHR])
    
    def create_department(self, request, *args, **kwargs):
        """
        Custom action to create a new department.
        Only Admin or HR can create a department.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        #self.perform_create(serializer)
        #headers = self.get_success_headers(serializer.data)
         # this will pass through company & name & employee_count,
        # and manager=None if not in request.data
        dept = serializer.save()

        return Response(
            DepartmentSerializer(dept).data,
            status=status.HTTP_201_CREATED
        )
 
    def get_queryset(self):
        qs =  super().get_queryset()
        u = self.request.user
        if u.role in ("Admin", "HR", "ADMIN"):
            return qs
        if u.role in ("HOD", "LM"):
            return qs.filter(manager=u)
        return qs.none() # regular employees cannot access departments
    
    def get_permissions(self):
        if self.action in ("list","retrieve"):
            return [IsAuthenticated()] # read-only for all authenticated users
        return [(IsAdmin | IsHR)()]  # write permissions for Admin & HR only

# ──────────────────────────────────────────────────────────────────────────────

class SubDepartmentViewSet(viewsets.ModelViewSet):
    queryset = SubDepartment.objects.select_related("department", "manager")
    serializer_class = SubDepartmentSerializer
    filter_backends  = [DjangoFilterBackend,filters.SearchFilter]
    filterset_fields = {
        "department": ["exact"],
        "department_id": ["exact"],
        "department__company": ["exact"],
        "department__company_id": ["exact"],
    }
    search_fields   = ["name", "department__name"]

    def get_permissions(self):
        # read-only for authenticated users, full access for Admin/HR
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [ReadOnlyOrAdminHR()]
        return [IsAuthenticated()]

class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.select_related("sub_department","manager")
    serializer_class = SectionSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = {
        "sub_department": ["exact"],                          # ?sub_department=<uuid>
        "sub_department_id": ["exact"],                       # ?sub_department_id=<uuid>
        "sub_department__department": ["exact"],              # ?sub_department__department=<uuid>
        "sub_department__department_id": ["exact"],           # ?sub_department__department_id=<uuid>
        "sub_department__department__company": ["exact"],     # ?sub_department__department__company=<uuid>
        "sub_department__department__company_id": ["exact"],  # ?sub_department__department__company_id=<uuid>
    }
    search_fields = ["name", "sub_department__name"]

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [ReadOnlyOrAdminHR()]
        return [IsAuthenticated()]

class SubSectionViewSet(viewsets.ModelViewSet):
    queryset = SubSection.objects.select_related("section","manager")
    serializer_class = SubSectionSerializer
    filter_backends = [DjangoFilterBackend,filters.SearchFilter]
    filterset_fields = {
        "section": ["exact"],                                      # ?section=<uuid>
        "section_id": ["exact"],                                   # ?section_id=<uuid>
        "section__sub_department": ["exact"],                      # ?section__sub_department=<uuid>
        "section__sub_department_id": ["exact"],                   # ?section__sub_department_id=<uuid>
        "section__sub_department__department": ["exact"],          # ?section__sub_department__department=<uuid>
        "section__sub_department__department_id": ["exact"],       # ?section__sub_department__department_id=<uuid>
        "section__sub_department__department__company_id": ["exact"],  # ?section__sub_department__department__company_id=<uuid>
    }
    search_fields = ["name","section__name"]

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [ReadOnlyOrAdminHR()]
        return [IsAuthenticated()]

class EmployeePlacementViewSet(viewsets.ModelViewSet):
    queryset = EmployeePlacement.objects.select_related(
        "employee","company","department","sub_department","section","sub_section",
        "department__manager","sub_department__manager","section__manager","sub_section__manager",
        "employee__user"
    )
    serializer_class = EmployeePlacementSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["employee__user__name","employee__user__email",
                     "department__name","sub_department__name","section__name","sub_section__name"]
    lookup_field = "employee_id"  # to resolve by employee_id from URL
    
    def get_permissions(self):
        # Admin/HR can write; everyone authenticated can read
        if self.action in ("create", "update", "partial_update", "destroy"):
            return [ReadOnlyOrAdminHR()]
        return [IsAuthenticated()]
    
    
    def get_queryset(self):
        u = self.request.user
        qs = super().get_queryset()
        if u.role in ("ADMIN","HR"):
            return qs
        if u.role in ("HOD","LM"):
            return qs.filter(
                Q(department__manager=u) |
                Q(sub_department__manager=u) |
                Q(section__manager=u) |
                Q(sub_section__manager=u)
            ).distinct()
        # employee: only his own placements
        return qs.filter(employee__user=u)
    
    def get_object(self):
        """
        Resolve by employee_id from URL and return that employee's single placement.
        """
        employee_id = self.kwargs[self.lookup_field]
        employee = get_object_or_404(Employee, employee_id=employee_id)
        # If you want 404 when not found:
        return get_object_or_404(EmployeePlacement, employee=employee)
    
    def partial_update(self, request, *args, **kwargs):
        # Normalize body ("" -> null), inject employee_id from URL
        data = request.data.copy()
        for k in ("department_id", "sub_department_id", "section_id", "sub_section_id"):
            if k in data and (data[k] == "" or data[k] is None):
                data[k] = None

        employee_id = self.kwargs[self.lookup_field]
        data["employee_id"] = employee_id

        instance = self.get_object()
        serializer = self.get_serializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)